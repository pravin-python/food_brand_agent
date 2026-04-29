"""
tools/report/report_tools.py
Excel report and heatmap builder tools for the Report Builder Agent
"""

import os
import json
from pathlib import Path
from collections import defaultdict

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

OUTPUT_DIR = Path("./output")
OUTPUT_DIR.mkdir(exist_ok=True)

# India state list for heatmap
INDIA_STATES = [
    "Andhra Pradesh", "Arunachal Pradesh", "Assam", "Bihar", "Chhattisgarh",
    "Goa", "Gujarat", "Haryana", "Himachal Pradesh", "Jharkhand", "Karnataka",
    "Kerala", "Madhya Pradesh", "Maharashtra", "Manipur", "Meghalaya", "Mizoram",
    "Nagaland", "Odisha", "Punjab", "Rajasthan", "Sikkim", "Tamil Nadu",
    "Telangana", "Tripura", "Uttar Pradesh", "Uttarakhand", "West Bengal",
    "Delhi", "Jammu and Kashmir", "Ladakh",
]


def build_excel(data: dict, filename: str) -> dict:
    """
    Build a multi-sheet Excel report from merged agent data.

    Args:
        data: Merged dict from orchestrator with all source data
        filename: Output filename (without path)

    Returns:
        dict with excel_path and status
    """
    if not HAS_OPENPYXL:
        return {"error": "openpyxl not installed. Run: pip install openpyxl"}

    brand      = data.get("brand", "Brand")
    filepath   = OUTPUT_DIR / f"{filename}.xlsx"
    wb         = openpyxl.Workbook()

    # ── Sheet 1: Summary ──────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Summary"
    _style_header(ws1, f"{brand} - India Presence Report")

    summary_rows = [
        ["Brand Name",       brand],
        ["Total States",     data.get("summary", {}).get("total_states", 0)],
        ["Total Cities",     data.get("summary", {}).get("total_cities", 0)],
        ["Confidence Score", f"{data.get('confidence', 0):.0%}"],
        ["Strongest State",  data.get("summary", {}).get("strongest_state", "")],
        ["Sources Used",     ", ".join(data.get("sources_used", []))],
    ]
    for i, (label, value) in enumerate(summary_rows, start=3):
        ws1.cell(row=i, column=1, value=label).font = Font(bold=True)
        ws1.cell(row=i, column=2, value=value)
    ws1.column_dimensions["A"].width = 20
    ws1.column_dimensions["B"].width = 35

    # ── Sheet 2: State-wise Presence ──────────────────────────────
    ws2 = wb.create_sheet("State Presence")
    headers = ["State", "Cities Found", "City List", "FSSAI", "E-comm", "Maps", "Confidence"]
    _write_header_row(ws2, headers)

    state_data = data.get("states_summary", {})
    for row_idx, (state, info) in enumerate(state_data.items(), start=2):
        cities     = info.get("cities", [])
        confidence = info.get("confidence", 0)
        fill_color = _confidence_color(confidence)
        row_data   = [
            state,
            len(cities),
            ", ".join(cities[:5]) + ("..." if len(cities) > 5 else ""),
            "Yes" if info.get("has_fssai")    else "No",
            "Yes" if info.get("has_ecommerce") else "No",
            "Yes" if info.get("has_maps")      else "No",
            f"{confidence:.0%}",
        ]
        for col_idx, val in enumerate(row_data, start=1):
            cell = ws2.cell(row=row_idx, column=col_idx, value=val)
            cell.fill = PatternFill(fill_type="solid", fgColor=fill_color)
    _autofit(ws2, headers)

    # ── Sheet 3: FSSAI Raw Data ────────────────────────────────────
    ws3      = wb.create_sheet("FSSAI Licenses")
    fssai_hd = ["License No", "Business Name", "State", "District", "City", "Address", "License Type"]
    _write_header_row(ws3, fssai_hd)
    for i, rec in enumerate(data.get("fssai_data", []), start=2):
        ws3.append([rec.get(k, "") for k in
                    ["license_no","business_name","state","district","city","address","license_type"]])
    _autofit(ws3, fssai_hd)

    # ── Sheet 4: E-commerce Availability ─────────────────────────
    ws4      = wb.create_sheet("E-commerce")
    ecomm    = data.get("ecommerce_data", {})
    platforms = ["swiggy", "blinkit", "amazon"]
    ecomm_hd  = ["City"] + [p.capitalize() for p in platforms] + ["Score"]
    _write_header_row(ws4, ecomm_hd)
    city_summary = ecomm.get("city_summary", {})
    for i, (city, info) in enumerate(city_summary.items(), start=2):
        avail = info.get("available_on", [])
        row   = [city] + [("Yes" if p in avail else "No") for p in platforms] + [info.get("score", 0)]
        ws4.append(row)
    _autofit(ws4, ecomm_hd)

    # ── Sheet 5: Distributor Locations ────────────────────────────
    ws5      = wb.create_sheet("Distributors")
    dist_hd  = ["Name", "City", "State", "Address", "Source", "Verified"]
    _write_header_row(ws5, dist_hd)
    for rec in data.get("maps_data", {}).get("distributors", []):
        ws5.append([rec.get(k, "") for k in
                    ["name","city","state","address","source","verified"]])
    _autofit(ws5, dist_hd)

    wb.save(filepath)
    return {"excel_path": str(filepath), "status": "success", "sheets": 5}


def build_heatmap(state_data: dict) -> dict:
    """
    Build heatmap data for India map visualization.

    Args:
        state_data: Dict of {state: confidence_score} (0.0-1.0)

    Returns:
        Dict with heatmap JSON data and intensity labels
    """
    heatmap = {}
    for state in INDIA_STATES:
        score = state_data.get(state, 0.0)
        heatmap[state] = {
            "score":     round(score, 2),
            "intensity": _score_to_label(score),
            "color":     _score_to_hex(score),
        }

    top_states = sorted(heatmap.items(), key=lambda x: x[1]["score"], reverse=True)

    return {
        "heatmap":    heatmap,
        "top_states": [s[0] for s in top_states[:5]],
        "legend":     {
            "Strong (>0.7)":   "#1D9E75",
            "Medium (0.4-0.7)": "#EF9F27",
            "Weak (<0.4)":     "#E24B4A",
            "Not found":       "#D3D1C7",
        },
    }


# ─── Helpers ─────────────────────────────────────────────────────

def _style_header(ws, title: str):
    ws["A1"] = title
    ws["A1"].font = Font(bold=True, size=14)
    ws["A1"].alignment = Alignment(horizontal="left")

def _write_header_row(ws, headers: list):
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(fill_type="solid", fgColor="534AB7")
        cell.alignment = Alignment(horizontal="center")

def _autofit(ws, headers: list):
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 18

def _confidence_color(score: float) -> str:
    if score >= 0.7:  return "C6EFCE"   # green
    if score >= 0.4:  return "FFEB9C"   # amber
    return "FFC7CE"                      # red

def _score_to_label(score: float) -> str:
    if score >= 0.7:  return "Strong"
    if score >= 0.4:  return "Medium"
    if score >  0.0:  return "Weak"
    return "Not found"

def _score_to_hex(score: float) -> str:
    if score >= 0.7:  return "#1D9E75"
    if score >= 0.4:  return "#EF9F27"
    if score >  0.0:  return "#E24B4A"
    return "#D3D1C7"


# ---- Register tools list for DeepAgent ----
REPORT_TOOLS = [build_excel, build_heatmap]
